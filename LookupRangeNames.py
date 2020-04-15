# LookupRangeNames.py
# mcyzyk, April 2020
#
# This script looks up Range Name per Call Number
#
# It takes as its input a spreadsheet consisting of a single column
# of call numbers, no header. This spreadsheet must be named "input.xlsx"
# and must be placed in the same directory with this script.
#
# It results in a CSV file, output.csv, consisting of two columns:
# One holding CallNumbers; the second holding associated RangeNames.
# No column headers.

from openpyxl import load_workbook
import sqlite3

print('Foo')

# Open spreadsheets
xlsx = load_workbook(filename="./input.xlsx")
xlsx = xlsx.active
xlsxLookup = load_workbook(filename="./LookupSheet.xlsx")
xlsxLookup = xlsxLookup.active

# Create in-memory database and lookup table in that database
DBLookupAndResults = sqlite3.connect("file::memory:?cache=shared")
cursor = DBLookupAndResults.cursor()
cursor.execute("DROP TABLE IF EXISTS lookup")
sql ='''CREATE TABLE lookup(
   RangeName CHAR(20) NOT NULL,
   BeginCallNumber CHAR(20),
   EndCallNumber CHAR(20),
   CollectionCode CHAR(20)
)'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("Lookup table created successfully........")

# Create results table in database
cursor.execute("DROP TABLE IF EXISTS results")
sql ='''CREATE TABLE results(
   CallNumber CHAR(20) NOT NULL,
   RangeName CHAR(20),
   NumberOfCheckouts INT,
   ItemNumber CHAR(20)
)'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("Results table created successfully........")

# Loop over Lookup sheet, inserting into Lookup database table
for row in xlsxLookup.iter_rows():
   cursor.execute('''INSERT INTO lookup(
   RangeName, BeginCallNumber, EndCallNumber, CollectionCode) VALUES
   ('str(row[0].value)', 'str(row[1].value)', 'str(row[1].value)', 'str(row[2].value)')''')
   print(str(row[0].value), str(row[1].value), str(row[2].value), str(row[3].value))
DBLookupAndResults.commit()
print("Lookup table populated successfully........")

# CollectionCodes:
# 	Eisenhower General Collection
# 	Eisenhower D Level Blue Labels
# 	Eisenhower A Level International Government Doc
# 	Eisenhower A Level Atlases

# dsAllRecords: Create dataset/temporary table of All Records, sorted
cursor.execute("DROP TABLE IF EXISTS dsAllRecords")
sql ='''CREATE TABLE dsAllRecords AS
   SELECT *
   FROM lookup
   ORDER BY BeginCallNumber
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsAllRecords dataset/temporary table created successfully........")

# dsEisenhowerGeneralCollection: Create dataset/temporary table of Eisenhower General Collection that are NOT Quarto or Folio
cursor.execute("DROP TABLE IF EXISTS dsEisenhowerGeneralCollection")
sql ='''CREATE TABLE dsEisenhowerGeneralCollection AS
   SELECT *
   FROM dsAllRecords
   WHERE
   BeginCallNumber NOT LIKE "%QUARTO" AND
   BeginCallNumber NOT LIKE "%FOLIO" AND
   CollectionCode = "Eisenhower General Collection"
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsEisenhowerGeneralCollection dataset/temporary table created successfully........")


# dsQuarto: Create dataset/temporary table of Quartos from Eisenhower General Collection
cursor.execute("DROP TABLE IF EXISTS dsQuarto")
sql ='''CREATE TABLE dsQuarto AS
   SELECT *
   FROM dsAllRecords
   WHERE
   BeginCallNumber LIKE "%QUARTO" AND
   CollectionCode = "Eisenhower General Collection"
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsQuarto dataset/temporary table created successfully........")

# dsFolio: Create dataset/temporary table of Folios from Eisenhower General Collection
cursor.execute("DROP TABLE IF EXISTS dsFolio")
sql ='''CREATE TABLE dsFolio AS
   SELECT *
   FROM dsAllRecords
   WHERE
   BeginCallNumber LIKE "%FOLIO" AND
   CollectionCode = "Eisenhower General Collection"
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsFolio dataset/temporary table created successfully........")

# dsEisenhowerDLevelBlueLabels: Create dataset/temporary table of Eisenhower D Level Blue Labels
cursor.execute("DROP TABLE IF EXISTS dsEisenhowerDLevelBlueLabels")
sql ='''CREATE TABLE dsEisenhowerDLevelBlueLabels AS
   SELECT *
   FROM dsAllRecords
   WHERE
   CollectionCode = "Eisenhower D Level Blue Labels"
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsEisenhowerDLevelBlueLabels dataset/temporary table created successfully........")

# dsEisenhowerALevelInternationalGovernmentDoc: Create dataset/temporary table of Eisenhower A Level International Government Doc
cursor.execute("DROP TABLE IF EXISTS dsEisenhowerALevelInternationalGovernmentDoc")
sql ='''CREATE TABLE dsEisenhowerALevelInternationalGovernmentDoc AS
   SELECT *
   FROM dsAllRecords
   WHERE
   CollectionCode = "Eisenhower A Level International Government Doc"
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsEisenhowerALevelInternationalGovernmentDoc dataset/temporary table created successfully........")

# dsEisenhowerALevelAtlases: Create dataset/temporary table of Eisenhower A Level Atlases
cursor.execute("DROP TABLE IF EXISTS dsEisenhowerALevelAtlases")
sql ='''CREATE TABLE dsEisenhowerALevelAtlases AS
   SELECT *
   FROM dsAllRecords
   WHERE
   CollectionCode = "Eisenhower A Level Atlases"
'''
cursor.execute(sql)
DBLookupAndResults.commit()
print("dsEisenhowerALevelAtlases dataset/temporary table created successfully........")
